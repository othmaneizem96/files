"""
app.py — RecrutAI Flask Web App (production)
=============================================
La clé API est cachée côté serveur via variable d'environnement.
Les utilisateurs ne voient jamais la clé.

Local  :  RESUMEPARSER_API_KEY=votre_cle  python app.py
Render :  définir RESUMEPARSER_API_KEY dans Environment Variables
"""

import os
import json
import uuid
import shutil
import io
from pathlib import Path
from datetime import datetime

from flask import (Flask, render_template, request, jsonify,
                   Response, send_file, stream_with_context)

from analyzer import analyze_all_stream

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ─── INIT ─────────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24))

# ── Clé API cachée — jamais envoyée au navigateur ──
API_KEY = os.environ.get("RESUMEPARSER_API_KEY", "")

UPLOAD_FOLDER = Path("uploads")
ALLOWED_EXTS  = {".pdf", ".docx", ".doc", ".txt"}
UPLOAD_FOLDER.mkdir(exist_ok=True)

# Résultats en mémoire par session
_sessions: dict[str, list] = {}


# ─── ROUTES ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    # On indique au template si la clé est configurée (sans l'exposer)
    key_configured = bool(API_KEY)
    return render_template("index.html", key_configured=key_configured)


@app.route("/upload", methods=["POST"])
def upload():
    sid    = request.form.get("session_id") or str(uuid.uuid4())
    folder = UPLOAD_FOLDER / sid
    folder.mkdir(parents=True, exist_ok=True)

    saved = []
    for f in request.files.getlist("cvs"):
        if Path(f.filename).suffix.lower() in ALLOWED_EXTS:
            dest = folder / f.filename
            f.save(dest)
            saved.append({"name": Path(f.filename).stem, "filename": f.filename})

    return jsonify({"session_id": sid, "uploaded": saved, "count": len(saved)})


@app.route("/analyze")
def analyze():
    """SSE — stream les résultats en temps réel. La clé API vient du serveur."""
    if not API_KEY:
        return jsonify({"error": "Clé API non configurée sur le serveur."}), 503

    sid       = request.args.get("session_id", "")
    job_title = request.args.get("job_title", "")
    job_desc  = request.args.get("job_desc", "")

    folder = UPLOAD_FOLDER / sid
    if not folder.exists():
        return jsonify({"error": "Session introuvable"}), 400

    cv_paths = [str(p) for p in folder.iterdir()
                if p.suffix.lower() in ALLOWED_EXTS]
    if not cv_paths:
        return jsonify({"error": "Aucun CV trouvé"}), 400

    _sessions[sid] = []

    def generate():
        yield f"data: {json.dumps({'type':'start','total':len(cv_paths)})}\n\n"

        for done, total, result in analyze_all_stream(
                cv_paths, API_KEY, job_title, job_desc):
            _sessions[sid].append(result)
            # On ne retourne PAS le solde API (info sensible)
            safe = {k: v for k, v in result.items() if k != "_balance"}
            yield f"data: {json.dumps({'type':'result','done':done,'total':total,'result':safe}, ensure_ascii=False)}\n\n"

        ok  = [r for r in _sessions[sid] if r.get("_statut") == "OK"]
        avg = round(sum(r.get("score_global", 0) for r in ok) / len(ok), 1) if ok else 0
        yield f"data: {json.dumps({'type':'done','ok':len(ok),'total':len(cv_paths),'avg':avg})}\n\n"

        # Nettoyage des fichiers uploadés après analyse
        _cleanup(sid)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


@app.route("/export/excel/<sid>")
def export_excel(sid):
    results   = _sessions.get(sid, [])
    ok        = sorted([r for r in results if r.get("_statut") == "OK"],
                       key=lambda x: x.get("score_global", 0), reverse=True)
    job_title = request.args.get("job_title", "Poste")

    wb  = _build_excel(ok, job_title)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"RecrutAI_{job_title.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/export/csv/<sid>")
def export_csv(sid):
    results = _sessions.get(sid, [])
    ok      = sorted([r for r in results if r.get("_statut") == "OK"],
                     key=lambda x: x.get("score_global", 0), reverse=True)

    lines = ["Rang,Nom,Score,Recommandation,Expérience,Niveau,Formation,Adéquation%,Email,Compétences"]
    for i, r in enumerate(ok, 1):
        skills = "|".join(r.get("competences_techniques", [])[:6])
        row    = [str(i), r.get("nom",""), str(r.get("score_global","")),
                  r.get("recommandation",""), str(r.get("experience_annees","")),
                  r.get("niveau",""), r.get("formation",""),
                  str(r.get("adequation_poste","")), r.get("email",""), skills]
        lines.append(",".join(f'"{v}"' for v in row))

    buf = io.BytesIO(("\ufeff" + "\n".join(lines)).encode("utf-8"))
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="RecrutAI_resultats.csv",
                     mimetype="text/csv")


@app.route("/status")
def status():
    """Indique si la clé API est configurée (sans l'exposer)."""
    return jsonify({"api_ready": bool(API_KEY)})


def _cleanup(sid: str):
    folder = UPLOAD_FOLDER / sid
    if folder.exists():
        shutil.rmtree(folder, ignore_errors=True)


# ─── EXCEL BUILDER ────────────────────────────────────────────────────────────

def _b():
    t = Side(style="thin", color="D4C4AE")
    return Border(left=t, right=t, top=t, bottom=t)

def _h(ws, row, col, val, bg="3D2B1A", fc="FAF7F2", bold=True, align="center", size=10):
    c = ws.cell(row=row, column=col)
    c.value = val
    c.font  = Font(name="Calibri", bold=bold, size=size, color=fc)
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _b()

def _d(ws, row, col, val, fc="1C1208", bg="FFFDF9", bold=False, align="left", wrap=False):
    c = ws.cell(row=row, column=col)
    c.value = val
    c.font  = Font(name="Calibri", size=10, color=fc, bold=bold)
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = _b()

def _sc(score):
    try:
        s = float(score)
        return ("D8F3DC","2D6A4F") if s>=8 else ("FFF3CD","856404") if s>=6 else ("FDE8E8","7D2226")
    except: return ("F3EDE3","6B5040")

def _rc(reco):
    r = str(reco).lower()
    return ("D8F3DC","2D6A4F") if "recommandé" in r else ("FFF3CD","856404") if "considérer" in r else ("FDE8E8","7D2226")

def _build_excel(results, job_title):
    wb = openpyxl.Workbook()

    # ── Feuille 1 : Tableau de bord ──
    ws1 = wb.active
    ws1.title = "📊 Tableau de bord"
    ws1.sheet_view.showGridLines = False
    for i, w in enumerate([2,5,26,20,12,10,10,10,24,2],1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.merge_cells("B1:J1"); ws1.row_dimensions[1].height = 44
    c = ws1["B1"]
    c.value = f"🎯  RecrutAI — {job_title}  |  {datetime.now().strftime('%d/%m/%Y')}"
    c.font  = Font(name="Calibri", bold=True, size=16, color="B07D4E")
    c.fill  = PatternFill("solid", fgColor="1C1208")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells("B2:J2"); ws1.row_dimensions[2].height = 18
    c2 = ws1["B2"]
    c2.value = f"{len(results)} candidats analysés via resumeparser.app"
    c2.font  = Font(name="Calibri", size=9, color="8A6A52")
    c2.fill  = PatternFill("solid", fgColor="2D1F10")
    c2.alignment = Alignment(horizontal="center", vertical="center")

    total = len(results)
    avg   = round(sum(r.get("score_global",0) for r in results)/total,1) if total else 0
    rec   = sum(1 for r in results if "recommandé" in str(r.get("recommandation","")).lower())
    cons  = sum(1 for r in results if "considérer" in str(r.get("recommandation","")).lower())
    nr    = sum(1 for r in results if "non retenu"  in str(r.get("recommandation","")).lower())

    ws1.row_dimensions[4].height = 14; ws1.row_dimensions[5].height = 36
    for (lbl,val,fc,bg),col in zip([
        ("Total CV",total,"3D2B1A","EDE4D6"), ("Score moyen",f"{avg}/10","3D2B1A","EDE4D6"),
        ("Recommandés",rec,"2D6A4F","D8F3DC"), ("À considérer",cons,"856404","FFF3CD"),
        ("Non retenus",nr,"7D2226","FDE8E8")
    ],[2,3,5,6,7]):
        lc = ws1.cell(row=4,column=col)
        lc.value=lbl; lc.font=Font(name="Calibri",size=8,color="6B5040")
        lc.alignment=Alignment(horizontal="center")
        vc = ws1.cell(row=5,column=col)
        vc.value=val; vc.font=Font(name="Calibri",bold=True,size=17,color=fc)
        vc.fill=PatternFill("solid",fgColor=bg)
        vc.alignment=Alignment(horizontal="center",vertical="center"); vc.border=_b()

    ws1.row_dimensions[7].height = 20
    for ci,h in enumerate(["#","Candidat","Poste","Exp.","Score","Adéq.","Niveau","Recommandation","Résumé"],2):
        _h(ws1,7,ci,h)

    medals={1:"🥇",2:"🥈",3:"🥉"}
    for ri,r in enumerate(results,8):
        ws1.row_dimensions[ri].height=32
        bg_row="FAF7F2" if ri%2==0 else "FFFDF9"
        score=r.get("score_global",0)
        sbg,sfc=_sc(score); rbg,rfc=_rc(r.get("recommandation",""))
        rank=ri-7
        _d(ws1,ri,2,medals.get(rank,rank),fc="B07D4E" if rank<=3 else "A08878",bg=bg_row,bold=rank<=3,align="center")
        _d(ws1,ri,3,r.get("nom",""),fc="1C1208",bg=bg_row,bold=rank<=3)
        _d(ws1,ri,4,r.get("poste_actuel",""),fc="5C4030",bg=bg_row)
        _d(ws1,ri,5,f"{r.get('experience_annees','?')} ans",fc="5C4030",bg=bg_row,align="center")
        _d(ws1,ri,6,f"{score}/10",fc=sfc,bg=sbg,bold=True,align="center")
        _d(ws1,ri,7,f"{r.get('adequation_poste',0)}%",fc="5C4030",bg=bg_row,align="center")
        _d(ws1,ri,8,r.get("niveau",""),fc="5C4030",bg=bg_row,align="center")
        _d(ws1,ri,9,r.get("recommandation",""),fc=rfc,bg=rbg,bold=True,align="center")
        _d(ws1,ri,10,str(r.get("resume_recruteur",""))[:100],fc="6B5040",bg=bg_row,wrap=True)
    ws1.freeze_panes="B8"

    # ── Feuille 2 : Fiches ──
    ws2 = wb.create_sheet("📋 Fiches détaillées")
    ws2.sheet_view.showGridLines=False
    ws2.column_dimensions["A"].width=2
    ws2.column_dimensions["B"].width=24
    ws2.column_dimensions["C"].width=42
    row=1
    for r in results:
        ws2.merge_cells(f"B{row}:C{row}"); ws2.row_dimensions[row].height=26
        c=ws2[f"B{row}"]
        c.value=f"👤  {r.get('nom',r.get('_fichier',''))}"
        c.font=Font(name="Calibri",bold=True,size=12,color="FAF7F2")
        c.fill=PatternFill("solid",fgColor="3D2B1A")
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        row+=1
        score=r.get("score_global",0); sbg,sfc=_sc(score); rbg,rfc=_rc(r.get("recommandation",""))
        fields=[
            ("Poste actuel",r.get("poste_actuel","—")),
            ("Expérience",f"{r.get('experience_annees','?')} ans — {r.get('niveau','?')}"),
            ("Formation",f"{r.get('formation','—')} {r.get('ecole','')}".strip()),
            ("Email",r.get("email","—")),("Téléphone",r.get("telephone","—")),
            ("Score global",f"{score}/10"),
            ("Compétences",", ".join(r.get("competences_techniques",[]))[:130] or "—"),
            ("Langues",", ".join(r.get("langues",[]))or"—"),
            ("Points forts"," | ".join(r.get("points_forts",[]))or"—"),
            ("Points faibles"," | ".join(r.get("points_faibles",[]))or"—"),
            ("Recommandation",r.get("recommandation","—")),
            ("Résumé",r.get("resume_recruteur","—")),
        ]
        for i,(lbl,val) in enumerate(fields):
            ws2.row_dimensions[row].height=18 if len(str(val))<90 else 30
            bg_l="F3EDE3" if i%2==0 else "FFFDF9"
            vbg,vfc=bg_l,"1C1208"
            if lbl=="Score global": vbg,vfc=sbg,sfc
            elif lbl=="Recommandation": vbg,vfc=rbg,rfc
            lc=ws2[f"B{row}"]
            lc.value=lbl; lc.font=Font(name="Calibri",bold=True,size=10,color="5C4030")
            lc.fill=PatternFill("solid",fgColor=bg_l)
            lc.alignment=Alignment(horizontal="left",vertical="center",indent=1); lc.border=_b()
            vc=ws2[f"C{row}"]
            vc.value=val; vc.font=Font(name="Calibri",size=10,color=vfc,bold=lbl=="Score global")
            vc.fill=PatternFill("solid",fgColor=vbg)
            vc.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1); vc.border=_b()
            row+=1
        row+=2

    # ── Feuille 3 : Classement ──
    ws3=wb.create_sheet("🏆 Classement")
    ws3.sheet_view.showGridLines=False
    for i,w in enumerate([2,7,26,18,10,10,10,14],1):
        ws3.column_dimensions[get_column_letter(i)].width=w
    ws3.merge_cells("B1:I1"); ws3.row_dimensions[1].height=28
    c=ws3["B1"]; c.value="🏆  Classement des candidats"
    c.font=Font(name="Calibri",bold=True,size=13,color="B07D4E")
    c.fill=PatternFill("solid",fgColor="1C1208")
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws3.row_dimensions[3].height=20
    for ci,h in enumerate(["Rang","Candidat","Poste","Score","Compét.","Exp.","Adéq.","Recommandation"],2):
        _h(ws3,3,ci,h)
    medals={1:"🥇",2:"🥈",3:"🥉"}
    for ri,r in enumerate(results,4):
        ws3.row_dimensions[ri].height=20; rank=ri-3
        score=r.get("score_global",0); sbg,sfc=_sc(score); rbg,rfc=_rc(r.get("recommandation",""))
        bg="FFFDF0" if rank<=3 else ("F3EDE3" if ri%2==0 else "FFFDF9")
        _d(ws3,ri,2,medals.get(rank,f"#{rank}"),fc="B07D4E" if rank<=3 else "A08878",bg=bg,bold=rank<=3,align="center")
        _d(ws3,ri,3,r.get("nom",""),fc="1C1208",bg=bg,bold=rank<=3)
        _d(ws3,ri,4,r.get("poste_actuel",""),fc="5C4030",bg=bg)
        _d(ws3,ri,5,f"{score}/10",fc=sfc,bg=sbg,bold=True,align="center")
        _d(ws3,ri,6,f"{r.get('score_competences','?')}/10",fc="5C4030",bg=bg,align="center")
        _d(ws3,ri,7,f"{r.get('experience_annees','?')} ans",fc="5C4030",bg=bg,align="center")
        _d(ws3,ri,8,f"{r.get('adequation_poste',0)}%",fc="5C4030",bg=bg,align="center")
        _d(ws3,ri,9,r.get("recommandation",""),fc=rfc,bg=rbg,bold=True,align="center")
    if results:
        chart=BarChart(); chart.type="bar"; chart.title="Score par candidat"; chart.style=10
        dr=Reference(ws3,min_col=5,min_row=4,max_row=3+len(results))
        cr=Reference(ws3,min_col=3,min_row=4,max_row=3+len(results))
        chart.add_data(dr); chart.set_categories(cr)
        chart.height=14; chart.width=28
        ws3.add_chart(chart,f"B{6+len(results)}")

    # ── Feuille 4 : Matrice ──
    ws4=wb.create_sheet("⚖️ Matrice compétences")
    ws4.sheet_view.showGridLines=False
    freq={}
    for r in results:
        for s in (r.get("competences_techniques") or []):
            k=s.lower().strip(); freq[k]=freq.get(k,0)+1
    top=sorted(freq,key=lambda x:-freq[x])[:22]; cands=results[:15]
    ws4.column_dimensions["A"].width=2; ws4.column_dimensions["B"].width=22
    for i in range(len(cands)): ws4.column_dimensions[get_column_letter(i+3)].width=13
    last_col=get_column_letter(2+len(cands)); ws4.merge_cells(f"B1:{last_col}1")
    ws4.row_dimensions[1].height=26
    c=ws4["B1"]; c.value="⚖️  Matrice compétences"
    c.font=Font(name="Calibri",bold=True,size=13,color="B07D4E")
    c.fill=PatternFill("solid",fgColor="1C1208"); c.alignment=Alignment(horizontal="center",vertical="center")
    ws4.row_dimensions[2].height=44; _h(ws4,2,2,"Compétence / Candidat")
    for ci,r in enumerate(cands,3):
        nom=r.get("nom",r.get("_fichier",""))
        c=ws4.cell(row=2,column=ci)
        c.value=f"{nom.split()[0] if nom else '?'}\n{r.get('score_global','?')}/10"
        c.font=Font(name="Calibri",bold=True,size=9,color="FAF7F2")
        c.fill=PatternFill("solid",fgColor="3D2B1A")
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=_b()
    for ro,skill in enumerate(top):
        row4=ro+3; ws4.row_dimensions[row4].height=17; bg_l="F3EDE3" if ro%2==0 else "FFFDF9"
        sc=ws4.cell(row=row4,column=2)
        sc.value=skill.title(); sc.font=Font(name="Calibri",bold=True,size=10,color="5C4030")
        sc.fill=PatternFill("solid",fgColor=bg_l)
        sc.alignment=Alignment(horizontal="left",vertical="center",indent=1); sc.border=_b()
        for ci,r in enumerate(cands,3):
            cskills=[s.lower() for s in (r.get("competences_techniques") or [])]
            has=any(skill in s or s in skill for s in cskills)
            cc=ws4.cell(row=row4,column=ci)
            cc.value="✓" if has else "—"
            cc.font=Font(name="Calibri",size=11,bold=has,color="2D6A4F" if has else "A08878")
            cc.fill=PatternFill("solid",fgColor="D8F3DC" if has else bg_l)
            cc.alignment=Alignment(horizontal="center",vertical="center"); cc.border=_b()

    return wb


# ─── RUN ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV") != "production"
    print(f"\n{'='*50}\n  🎯 RecrutAI — Flask Web App\n  📍 http://localhost:{port}\n{'='*50}\n")
    app.run(host="0.0.0.0", port=port, debug=debug, threaded=True)
